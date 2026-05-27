# excel_reporter.py
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def exportar_a_excel_ejecutivo(df_vulnerables: pd.DataFrame, ruta_salida: str = "data/informe_servidores_criticos.xlsx") -> None:
    """
    Toma el DataFrame filtrado de servidores y genera un archivo Excel pulido
    y formateado profesionalmente para la gerencia.
    """
    import os
    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)

    print(f"📊 Generando reporte ejecutivo en Excel: {ruta_salida}...")
    
    # 1. Exportación directa desde Pandas
    df_vulnerables.to_excel(ruta_salida, sheet_name="Servidores Críticos", index=False)
    
    # 2. Estilizado avanzado con OpenPyXL para que luzca corporativo
    wb = load_workbook(ruta_salida)
    ws = wb["Servidores Críticos"]
    
    # Paleta de colores: Azul oscuro desaturado para administración
    color_cabecera = "1F497D"
    fuente_cabecera = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    relleno_cabecera = PatternFill(start_color=color_cabecera, end_color=color_cabecera, fill_type="solid")
    
    fuente_datos = Font(name="Arial", size=10, bold=False)
    alineacion_centro = Alignment(horizontal="center", vertical="center")
    alineacion_izquierda = Alignment(horizontal="left", vertical="center")
    
    borde_fino = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Aplicar estilos a la cabecera
    for col in range(1, ws.max_column + 1):
        celda = ws.cell(row=1, column=col)
        celda.font = fuente_cabecera
        celda.fill = relleno_cabecera
        celda.alignment = alineacion_centro
    
    # Aplicar estilos a las filas de datos y ajustar alineaciones
    for fila in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            celda = ws.cell(row=fila, column=col)
            celda.font = fuente_datos
            celda.border = borde_fino
            
            # Formatear columnas específicas
            if ws.cell(row=1, column=col).value in ["IP", "RAM_GB"]:
                celda.alignment = alineacion_centro
            else:
                celda.alignment = alineacion_izquierda
                
            # Si es la columna de RAM, añadir sufijo para que sea legible
            if ws.cell(row=1, column=col).value == "RAM_GB":
                celda.value = f"{celda.value} GB"

    # Autoajustar el ancho de las columnas para evitar desbordamientos o texto cortado
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Guardar los cambios finales sobre el archivo
    wb.save(ruta_salida)
    print("✅ ¡Reporte Excel formateado y guardado correctamente!")

if __name__ == "__main__":
    # Prueba rápida de generación autónoma
    try:
        from inventory_manager import procesar_inventario
        df_criticos, _ = procesar_inventario("inventory.csv")
        exportar_a_excel_ejecutivo(df_criticos)
    except Exception as e:
        print(f"❌ Ejecuta primero 'generate_inventory.py' para probar este módulo. Error: {e}")